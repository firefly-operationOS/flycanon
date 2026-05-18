# Copyright 2026 Firefly Software Solutions Inc
"""``MetadataExtractor`` -- per-format metadata extraction.

For every supported format the extractor reads the embedded metadata
(author, title, creation date, keywords) directly from the bytes,
detects the dominant language of the extracted content, and
normalises the result into a flat :class:`ExtractedMetadata` dict.
The :class:`IntakeService` merges this dict into
``SourceRow.metadata_json`` so retrieval-time filters
(``metadata.language``, ``metadata.author``, ``metadata.pages``)
become first-class facets.

Format coverage:

* PDF   -- PyMuPDF ``doc.metadata`` + page count.
* DOCX  -- python-docx ``core_properties``.
* PPTX  -- python-pptx ``core_properties`` + slide count.
* XLSX  -- openpyxl ``workbook.properties`` + sheet names.
* HTML  -- BeautifulSoup ``<title>`` + ``<meta name=...>``.
* Email -- stdlib ``email`` headers (subject / from / date).
* Plain text / Markdown / JSON / XML -- language detection only.
* Images -- PIL EXIF (camera, orientation, capture date).

Everything is best-effort: a missing/corrupt header never aborts the
ingest -- the field is simply omitted from ``extracted``.
"""

from __future__ import annotations

import io
import logging
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any

from pyfly.container import service

logger = logging.getLogger(__name__)


@dataclass(slots=True)
class ExtractedMetadata:
    """Normalised metadata fields the extractor emits.

    All fields are optional; the extractor returns whichever the
    document carries. Dates are ISO-8601 strings rather than
    :class:`datetime` so the dict serialises cleanly via JSON
    columns without bespoke encoders.
    """

    title: str | None = None
    author: str | None = None
    subject: str | None = None
    keywords: list[str] = field(default_factory=list)
    language: str | None = None
    created_at: str | None = None
    modified_at: str | None = None
    pages: int | None = None
    word_count: int | None = None
    extra: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        out: dict[str, Any] = {}
        for k, v in (
            ("title", self.title),
            ("author", self.author),
            ("subject", self.subject),
            ("language", self.language),
            ("created_at", self.created_at),
            ("modified_at", self.modified_at),
            ("pages", self.pages),
            ("word_count", self.word_count),
        ):
            if v not in (None, ""):
                out[k] = v
        if self.keywords:
            out["keywords"] = self.keywords
        if self.extra:
            out["extra"] = self.extra
        return out


@service
class MetadataExtractor:
    """Per-format metadata extractor."""

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def extract(
        self,
        data: bytes,
        *,
        media_type: str,
        filename: str | None = None,
        text_sample: str | None = None,
    ) -> ExtractedMetadata:
        """Return the document's metadata. Never raises -- returns
        an empty :class:`ExtractedMetadata` on any failure.
        """
        try:
            if media_type == "application/pdf":
                return self._extract_pdf(data, text_sample=text_sample)
            if media_type == (
                "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            ):
                return self._extract_docx(data, text_sample=text_sample)
            if media_type == (
                "application/vnd.openxmlformats-officedocument.presentationml.presentation"
            ):
                return self._extract_pptx(data, text_sample=text_sample)
            if media_type == (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ):
                return self._extract_xlsx(data, text_sample=text_sample)
            if media_type in {"text/html", "application/xhtml+xml"}:
                return self._extract_html(data, text_sample=text_sample)
            if media_type == "message/rfc822":
                return self._extract_eml(data, text_sample=text_sample)
            if media_type.startswith("image/"):
                return self._extract_image(data, text_sample=text_sample)
            # Plain text / Markdown / JSON / XML / CSV / transcripts:
            # nothing structural to extract, but the language detector
            # still gives us a useful facet.
            return self._language_only(text_sample)
        except Exception as exc:
            logger.debug(
                "metadata extraction failed media=%s filename=%s err=%s",
                media_type,
                filename,
                exc,
            )
            return self._language_only(text_sample)

    # ------------------------------------------------------------------
    # Format-specific extractors
    # ------------------------------------------------------------------

    def _extract_pdf(self, data: bytes, *, text_sample: str | None) -> ExtractedMetadata:
        import pymupdf

        with pymupdf.open(stream=data, filetype="pdf") as doc:
            meta = dict(doc.metadata or {})
            pages = doc.page_count
        return ExtractedMetadata(
            title=_clean(meta.get("title")),
            author=_clean(meta.get("author")),
            subject=_clean(meta.get("subject")),
            keywords=_split_keywords(meta.get("keywords")),
            language=_detect_language(text_sample),
            created_at=_iso(meta.get("creationDate")),
            modified_at=_iso(meta.get("modDate")),
            pages=pages,
            extra={k: meta[k] for k in ("producer", "creator") if meta.get(k)},
        )

    def _extract_docx(self, data: bytes, *, text_sample: str | None) -> ExtractedMetadata:
        from docx import Document

        doc = Document(io.BytesIO(data))
        cp = doc.core_properties
        word_count = sum(len(p.text.split()) for p in doc.paragraphs)
        return ExtractedMetadata(
            title=_clean(cp.title),
            author=_clean(cp.author),
            subject=_clean(cp.subject),
            keywords=_split_keywords(cp.keywords),
            language=_clean(cp.language) or _detect_language(text_sample),
            created_at=_iso(cp.created),
            modified_at=_iso(cp.modified),
            word_count=word_count or None,
            extra=_kv("last_modified_by", cp.last_modified_by, "revision", cp.revision),
        )

    def _extract_pptx(self, data: bytes, *, text_sample: str | None) -> ExtractedMetadata:
        from pptx import Presentation

        prs = Presentation(io.BytesIO(data))
        cp = prs.core_properties
        return ExtractedMetadata(
            title=_clean(cp.title),
            author=_clean(cp.author),
            subject=_clean(cp.subject),
            keywords=_split_keywords(cp.keywords),
            language=_clean(cp.language) or _detect_language(text_sample),
            created_at=_iso(cp.created),
            modified_at=_iso(cp.modified),
            pages=len(prs.slides),
            extra=_kv("last_modified_by", cp.last_modified_by),
        )

    def _extract_xlsx(self, data: bytes, *, text_sample: str | None) -> ExtractedMetadata:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        props = wb.properties
        sheet_names = list(wb.sheetnames)
        try:
            return ExtractedMetadata(
                title=_clean(props.title),
                author=_clean(props.creator),
                subject=_clean(props.subject),
                keywords=_split_keywords(props.keywords),
                language=_clean(props.language) or _detect_language(text_sample),
                created_at=_iso(props.created),
                modified_at=_iso(props.modified),
                pages=len(sheet_names),
                extra=_kv("last_modified_by", props.lastModifiedBy, "sheets", sheet_names),
            )
        finally:
            wb.close()

    def _extract_html(self, data: bytes, *, text_sample: str | None) -> ExtractedMetadata:
        from bs4 import BeautifulSoup

        soup = BeautifulSoup(data, "lxml")
        title = soup.title.string.strip() if soup.title and soup.title.string else None
        meta_extra: dict[str, Any] = {}
        keywords: list[str] = []
        author: str | None = None
        description: str | None = None
        language: str | None = None
        for tag in soup.find_all("meta"):
            name = (tag.get("name") or tag.get("property") or "").lower()
            content = tag.get("content")
            if not (name and content):
                continue
            if name in {"author"}:
                author = content
            elif name in {"description", "og:description"}:
                description = content
            elif name == "keywords":
                keywords = _split_keywords(content)
            elif name in {"language", "content-language"}:
                language = content
            else:
                meta_extra[name] = content
        return ExtractedMetadata(
            title=_clean(title),
            author=_clean(author),
            subject=_clean(description),
            keywords=keywords,
            language=_clean(language) or _detect_language(text_sample),
            extra=meta_extra,
        )

    def _extract_eml(self, data: bytes, *, text_sample: str | None) -> ExtractedMetadata:
        import email
        from email.policy import default as default_policy

        msg = email.message_from_bytes(data, policy=default_policy)
        return ExtractedMetadata(
            title=_clean(msg.get("Subject")),
            author=_clean(msg.get("From")),
            subject=_clean(msg.get("Subject")),
            created_at=_iso(msg.get("Date")),
            language=_detect_language(text_sample),
            extra=_kv(
                "to", msg.get("To"),
                "cc", msg.get("Cc"),
                "message_id", msg.get("Message-ID"),
            ),
        )

    def _extract_image(self, data: bytes, *, text_sample: str | None) -> ExtractedMetadata:
        from PIL import ExifTags, Image

        try:
            with Image.open(io.BytesIO(data)) as img:
                width, height = img.size
                exif = img.getexif() or {}
        except Exception:
            return ExtractedMetadata(language=_detect_language(text_sample))

        tags = {ExifTags.TAGS.get(k, str(k)): v for k, v in exif.items()}
        date = tags.get("DateTimeOriginal") or tags.get("DateTime")
        return ExtractedMetadata(
            title=tags.get("ImageDescription") or None,
            author=tags.get("Artist") or None,
            created_at=_clean(date),
            language=_detect_language(text_sample),
            extra={
                k: v for k, v in {
                    "make": tags.get("Make"),
                    "model": tags.get("Model"),
                    "orientation": tags.get("Orientation"),
                    "width": width,
                    "height": height,
                }.items()
                if v not in (None, "")
            },
        )

    def _language_only(self, text_sample: str | None) -> ExtractedMetadata:
        return ExtractedMetadata(language=_detect_language(text_sample))


# ----------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _split_keywords(value: Any) -> list[str]:
    if not value:
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    s = str(value)
    parts = [p.strip() for p in s.replace(";", ",").split(",")]
    return [p for p in parts if p]


def _iso(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    s = str(value).strip()
    if not s:
        return None
    # PDF date format: ``D:YYYYMMDDHHMMSS+TZ``.
    if s.startswith("D:"):
        s = s[2:]
        try:
            return datetime.strptime(s[:14], "%Y%m%d%H%M%S").isoformat()
        except ValueError:
            return s
    return s


def _kv(*pairs: Any) -> dict[str, Any]:
    """Build a dict from interleaved ``key, value`` args, dropping
    entries whose value is falsy/None.
    """
    it = iter(pairs)
    out: dict[str, Any] = {}
    for key, val in zip(it, it, strict=False):
        if val not in (None, "", [], {}):
            out[str(key)] = val
    return out


def _detect_language(text: str | None) -> str | None:
    if not text or len(text) < 40:
        return None
    try:
        from langdetect import DetectorFactory, detect

        DetectorFactory.seed = 0  # deterministic per process
        return detect(text)
    except Exception:
        return None
